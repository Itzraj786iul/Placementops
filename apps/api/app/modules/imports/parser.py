"""Parse CSV / XLSX shortlist files into identifier rows."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from openpyxl import load_workbook

from app.modules.imports.enums import HEADER_ALIASES, MatchField
from app.modules.imports.exceptions import ImportValidationError


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    identifier: str


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _find_column_index(headers: list[str], match_field: MatchField) -> int:
    aliases = HEADER_ALIASES[match_field]
    for idx, header in enumerate(headers):
        if header in aliases:
            return idx
    # Fallback: single-column file
    if len(headers) == 1 and headers[0]:
        return 0
    raise ImportValidationError(
        f"Could not find a {match_field.value.replace('_', ' ').title()} column. "
        f"Expected one of: {', '.join(sorted(aliases))}"
    )


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_csv(content: bytes, match_field: MatchField) -> list[ParsedRow]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportValidationError("CSV must be UTF-8 encoded") from exc

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ImportValidationError("File is empty")

    headers = [_normalize_header(h) for h in rows[0]]
    col = _find_column_index(headers, match_field)

    parsed: list[ParsedRow] = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or all(not str(c).strip() for c in row):
            continue
        identifier = _cell_str(row[col] if col < len(row) else "")
        parsed.append(ParsedRow(row_number=i, identifier=identifier))
    if not parsed:
        raise ImportValidationError("No data rows found in file")
    return parsed


def parse_xlsx(content: bytes, match_field: MatchField) -> list[ParsedRow]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError("Invalid XLSX file") from exc

    ws = wb.active
    if ws is None:
        raise ImportValidationError("Workbook has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ImportValidationError("File is empty") from exc

    headers = [_normalize_header(h) for h in header_row]
    col = _find_column_index(headers, match_field)

    parsed: list[ParsedRow] = []
    for i, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        identifier = _cell_str(row[col] if col < len(row) else "")
        parsed.append(ParsedRow(row_number=i, identifier=identifier))

    wb.close()
    if not parsed:
        raise ImportValidationError("No data rows found in file")
    return parsed


def parse_shortlist_file(
    filename: str,
    content: bytes,
    match_field: MatchField,
) -> list[ParsedRow]:
    if not content:
        raise ImportValidationError("Uploaded file is empty")

    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content, match_field)
    if lower.endswith(".xlsx"):
        return parse_xlsx(content, match_field)
    raise ImportValidationError("Unsupported file type. Upload CSV or XLSX.")
